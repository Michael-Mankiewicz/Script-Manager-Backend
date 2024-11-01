import csv
import os
from datetime import datetime
import shutil
import pandas as pd
from openpyxl.styles import Font, Border, PatternFill, Alignment, Side
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.cell.text import RichText
from io import BytesIO
import tempfile
from PIL import Image as PILImage
from django.conf import settings

class AddressCorrectionBot:
    def __init__(self, carton_file, fedex_invoice):
        self.carton_file = carton_file
        self.fedex_invoice = fedex_invoice
        self.ADDRESS_CORRECTION_FEE = 22.50

    def process_files(self):

        resources_path = os.path.join(settings.MEDIA_ROOT, 'resources')
        results_path = os.path.join(settings.MEDIA_ROOT, 'results')

        if not os.path.exists(resources_path):
            os.makedirs(resources_path)

        if not os.path.exists(results_path):
            os.makedirs(results_path)

        self.clear_results_directory(results_path)

        abs_cartonfile_path = os.path.join(settings.MEDIA_ROOT, self.carton_file)
        abs_fedexinvoice_path = os.path.join(settings.MEDIA_ROOT, self.fedex_invoice)
        print(f"Absolute Carton File Path: {abs_cartonfile_path}")
        print(f"Absolute FedEx Invoice Path: {abs_fedexinvoice_path}")
        if not os.path.isfile(abs_cartonfile_path):
            raise FileNotFoundError(f"Carton file not found at {abs_cartonfile_path}")
        if not os.path.isfile(abs_fedexinvoice_path):
            raise FileNotFoundError(f"FedEx invoice file not found at {abs_fedexinvoice_path}")
        
        print("type of invoice path:" + str(type(abs_fedexinvoice_path)))
        fedexInvoiceRowsWithAddrCorr = self.AddressCorrectionSearch(abs_fedexinvoice_path)
        combinedInfo = self.CompileNewCSVFile(fedexInvoiceRowsWithAddrCorr, abs_cartonfile_path)
        projectsData = self.SortCSVbyProject(combinedInfo)
        self.CreateInvoice(projectsData)
        results_dir = os.path.join(settings.MEDIA_ROOT, 'results')
        os.makedirs(results_dir, exist_ok=True)
        self.delete_non_xlsx_files(results_dir)

        generated_files = [os.path.join(results_path, filename) for filename in os.listdir(results_path)]
        return generated_files


    def find_address_correction_column(self, csv_file_path, search_term):
        with open(csv_file_path, 'r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            for row_num, row in enumerate(reader):
                for col_index, cell in enumerate(row):
                    #print("read this cell: " + cell + " at " + str(col_index))
                    if search_term == cell:
                        return col_index, row
        return None, None

    def AddressCorrectionSearch(self, csvFilePath):
        search_term = 'Address Correction'
        resulting_rows = []
        
        with open(csvFilePath, 'r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            header = next(reader)
            col_index, addr_corr_row = self.find_address_correction_column(csvFilePath, search_term)
            if col_index == None:
                raise ValueError(f"No rows matching {search_term} were found")

            for row_num, row in enumerate(reader):
                if len(row) > col_index:  # Ensure the row has enough columns
                    cell_content = row[col_index].strip()  # Use .strip() to remove any leading/trailing spaces
                    if cell_content == search_term:
                        resulting_rows.append(row)
                else:
                    print(f"Row {row_num} does not have a column {col_index}")

        #print("Rows after filter:", len(resulting_rows))
        return resulting_rows


    def CompileNewCSVFile(self, addrCorrRows, cartonFilePath):
        fedexInvoiceColumns = [
        1,   # Invoice Date
        2,   # Invoice Number
        8,   # Express or Ground Tracking Number
        13,  # Shipment Date
        32,  # Recipient Name
        34,  # Recipient Address Line 1
        35,  # Recipient Address Line 2
        36,  # Recipient City
        37,  # Recipient State
        38,  # Recipient Zip Code
        39,  # Recipient Country/Territory
        48,  # Original Customer Reference
        49,  # Original Ref#2
        50,  # Original Ref #3/PO Number
        108  # Tracking ID Charge Description
        ]

        addrCorrInvoiceData = []

        current_date = datetime.now().strftime("%Y-%m-%d")

        for row in addrCorrRows:
            infoNeededFromFedex = []
            for index in fedexInvoiceColumns:
                if index <= len(row): 
                    infoNeededFromFedex.append(row[index])

            addrCorrInvoiceData.append(infoNeededFromFedex)

        self.LinkProject(addrCorrInvoiceData, cartonFilePath)

        results_dir = os.path.join(settings.MEDIA_ROOT, 'results')
        os.makedirs(results_dir, exist_ok=True)
        unsortedAddressCorrectionCSV = os.path.join(results_dir, f'Ungrouped_projects_{current_date}.csv')

        with open(unsortedAddressCorrectionCSV, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            for row in addrCorrInvoiceData:
                writer.writerow(row)
        return(addrCorrInvoiceData)
    
        

    def LinkProject(self,addrCorrRows, cartonFilePath):
        cartonFileTrackingNumIndex = 5 
        fedexInvoiceTrackingNumIndex = 2
        cartonFileProjectIndex = 0
        cartonFileOwnerReferenceIndex = 19

        # Read the new CSV file and store it in a dictionary
        carton_dict = {}
        with open(cartonFilePath, 'r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            for row in reader:
                if len(row) > cartonFileTrackingNumIndex:
                    # Store the row in the dictionary where the key is the tracking number and the entire row is the value
                    cartonFileTrackingNum = row[cartonFileTrackingNumIndex].strip()
                    carton_dict[cartonFileTrackingNum] = row 

        #add project and owner reference from cartonfile to addrcorrectinfo array
        for row in addrCorrRows:
            trackingNumber = row[fedexInvoiceTrackingNumIndex].strip()
            if trackingNumber in carton_dict:
                row.append(carton_dict[trackingNumber][cartonFileProjectIndex])
                row.append(carton_dict[trackingNumber][cartonFileOwnerReferenceIndex])
            else:
                print("Key not found in carton file:", trackingNumber)  # Debugging line

        print("Total Rows Updated in rowsList:", len(addrCorrRows))  # Debugging line

    def SortCSVbyProject(self, addrCorrList):
        project_col_index = 15  # Column 17 (0-indexed)
        backup_project_col_index = 11  # Column 12 (0-indexed)
        grouped_projects = ['Cosmedix', 'Pur', 'Butter London', 'Aloette']
        grouped_project_name = 'Grouped_Project'
        projects = {}

        current_date = datetime.now().strftime("%Y-%m-%d")
        results_dir = os.path.join(settings.MEDIA_ROOT, 'results')
        os.makedirs(results_dir, exist_ok=True)

        for row in addrCorrList:
            # Ensure row has enough columns
            while len(row) <= project_col_index:
                row.append('')

            project_name = row[project_col_index].strip()

            if not project_name and 'fab' in row[backup_project_col_index].lower():
                project_name = 'First Aid Beauty'
                row[project_col_index] = 'First Aid Beauty'

            if project_name in grouped_projects:
                project_name = grouped_project_name

            if project_name:
                if project_name not in projects:
                    projects[project_name] = []
                projects[project_name].append(row)
            else:
                print("Row without a project:", row)

        for project_name, rows in projects.items():
            filename = os.path.join(results_dir, f'{project_name}_project_{current_date}.csv')
            with open(filename, 'w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerows(rows)

        print(f"Created {len(projects)} project CSV files.")

        return projects

    def CreateInvoice(self, projectsData):
        # Generate an invoice based on a CSV file
        
        current_date = datetime.now().strftime("%Y-%m-%d")

        regular_font = Font(bold=False)  # Define a regular font style
        bold_font = Font(bold=True)  # Define a bold font style
        bold_underline_font = Font(bold=True, underline='single')  # Define a bold and underlined font style
        gray_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')  # Define a gray fill


        def copy_header(source_sheet, dest_sheet):
            # Copy header (rows 1-6)
            for row in range(1, 7):
                for col in range(1, 8):
                    source_cell = source_sheet.cell(row=row, column=col)
                    dest_cell = dest_sheet.cell(row=row, column=col)
                    dest_cell.value = source_cell.value
                    copy_cell_style(source_cell, dest_cell)

        def copy_cell_style(src_cell, dest_cell):
            # Copy font properties
            dest_font = Font(name=src_cell.font.name, size=src_cell.font.size, bold=src_cell.font.bold, italic=src_cell.font.italic,
                            vertAlign=src_cell.font.vertAlign, underline=src_cell.font.underline, strike=src_cell.font.strike,
                            color=src_cell.font.color)
            dest_cell.font = dest_font

            # Remove cell borders
            #dest_cell.border = Border()

            # Copy fill properties
            dest_fill = PatternFill(fill_type=src_cell.fill.fill_type, start_color=src_cell.fill.start_color,
                                    end_color=src_cell.fill.end_color)
            dest_cell.fill = dest_fill

            # Copy alignment and number format
            dest_cell.alignment = Alignment(horizontal=src_cell.alignment.horizontal, vertical=src_cell.alignment.vertical)
            dest_cell.number_format = src_cell.number_format

        def copy_range_with_style(start_col, start_row, end_col, end_row, source_sheet, dest_sheet):
            for i in range(start_row, end_row + 1):
                for j in range(start_col, end_col + 1):
                    source_cell = source_sheet.cell(row=i, column=j)
                    dest_cell = dest_sheet.cell(row=i, column=j)
                    dest_cell.value = source_cell.value
                    copy_cell_style(source_cell, dest_cell)

        def copy_dimensions(source_sheet, dest_sheet):
            # Copy row dimensions
            for row in source_sheet.row_dimensions.values():
                dest_sheet.row_dimensions[row.index].height = row.height

            # Copy column dimensions
            for col in source_sheet.column_dimensions.values():
                dest_sheet.column_dimensions[col.index].width = col.width

        def copy_merged_cells(source_sheet, dest_sheet):
            # Copy merged cell ranges
            for merged_cell_range in source_sheet.merged_cells.ranges:
                dest_sheet.merge_cells(str(merged_cell_range))

        # Load your Excel template
        resources_dir = os.path.join(settings.MEDIA_ROOT, 'resources')
        os.makedirs(resources_dir, exist_ok=True)
        template_path = os.path.join(resources_dir, 'invoice_template.xlsx')

        source_wb = load_workbook(template_path)
        source_ws = source_wb.worksheets[0]  # Use the first sheet in the workbook

        for project_name, data_lists in projectsData.items():
            # Create a new workbook for each project
            wb = Workbook()
            ws = wb.active
            ws.title = project_name

            # Copy header
            copy_header(source_ws, ws)


            # Copy range with style
            copy_range_with_style(1, 1, 7, 13, source_ws, ws)

            # Copy cell dimensions
            copy_dimensions(source_ws, ws)

            # Copy merged cells
            copy_merged_cells(source_ws, ws)
            
            total = 0

            # Temporary directory to store images
            #temp_dir = tempfile.mkdtemp()

            # Copy PNG image


            image_path = os.path.join(resources_dir, 'fedexLogo.png')
            desired_size = (180, 50)
            # Open the image using PIL
            original_img = PILImage.open(image_path)

            # Resize the image
            resized_img = original_img.resize(desired_size)

            # Save the resized image to a BytesIO object
            image_stream = BytesIO()
            resized_img.save(image_stream, format='PNG')
            image_stream.seek(0)

            # Load the image into openpyxl
            img = Image(image_stream)
            img.anchor = 'A3'
            ws.add_image(img)

            #add project name underneath "Project"
            ws.cell(row=4, column=3).value = project_name

            # Loop through data and create invoice blocks
            for index, data_list in enumerate(data_lists):
                start_row = 7 + (index * 7)  # New block every 7 rows
                total += self.ADDRESS_CORRECTION_FEE

                #print(f"data_list[15] before conversion: '{data_list[15]}'")

                # Enter data into specific cells
                if len(data_list) > 6:
                    date_obj = datetime.strptime(data_list[0], '%Y%m%d').strftime('%b %d, %Y')

                    shipdate_cell = ws.cell(row=start_row, column=1)
                    
                    shipdate_cell.value = f'Invoice Date: {date_obj}'
                    shipdate_cell.font = Font(bold=True)# Ship Date
                    date_obj = datetime.strptime(data_list[3], '%Y%m%d').strftime('%b %d, %Y')
                    ws.cell(row=start_row + 4, column=1).value = 'Ship Date: '  # Delivered Label
                    ws.cell(row=start_row + 4, column=2).value = date_obj  # Delivered Var

                    #payor
                    payorCell = ws.cell(row=start_row + 1, column=1)
                    payorValue = 'Payor: Shipper'
                    payorCell.value = payorValue
                    payorCell.font = Font(bold=True)
                    ws.cell(row=start_row + 1, column=1).value = 'Payor: Shipper'

                    ws.cell(row=start_row + 2, column=1).value = 'Tracking ID'  # Tracking ID Label

                    #Customer Reference
                    custRefCell =  ws.cell(row=start_row, column=3)
                    custRefValue = f'Cust. Ref: {data_list[11]}'
                    custRefCell.value = custRefValue
                    custRefCell.font = Font(bold=True)
                    #ws.cell(row=start_row, column=3).value = str('Cust. Ref.: ' + data_list[11]) 
                    
                    # Dept
                    deptCell = ws.cell(row=start_row + 1, column=3)
                    deptValue = f'Dept.#: {data_list[12]}'
                    deptCell.value = deptValue
                    deptCell.font = Font(bold=True)
                    #ws.cell(row=start_row + 1, column=3).value = str('Dept.#: ' + data_list[12])  

                    #PO
                    poCell = ws.cell(row=start_row, column=5)
                    poCellValue = f'P.O.#: {data_list[13]}'
                    poCell.value = poCellValue
                    poCell.font = Font(bold=True)

                    #ws.cell(row=start_row, column=5).value = str('P.O.#: ' + data_list[13])  # PO


                    #owner ref
                    try:
                        ownerRefCell =  ws.cell(row=start_row + 1, column=5)
                        ownerRefValue = f'Owner Ref.: {data_list[17]}'
                        ownerRefCell.value = ownerRefValue
                        ownerRefCell.font = Font(bold=True)
                    # ws.cell(row=start_row + 1, column=5).value = 'Owner Ref.: ' + str(data_list[17])
                    except IndexError:
                        ownerRefCell =  ws.cell(row=start_row + 1, column=5)
                        ownerRefValue = ""
                        ownerRefCell.value = ownerRefValue
                        ownerRefCell.font = Font(bold=True)
                        #ws.cell(row=start_row + 1, column=5).value = 'Owner Ref.: '  # or some default value
                    
                    recipient_cell = ws.cell(row=start_row + 2, column=3)
                    recipient_cell.value = 'Recipient'
                    recipient_cell.font = Font(bold=True, underline='single')
                    
                    ws.cell(row=start_row + 2, column=2).value = data_list[2]  # 
                    ws.cell(row=start_row + 3, column=3).value = data_list[4]  # 
                    ws.cell(row=start_row + 4, column=3).value = str(data_list[5] + ' ' + data_list[6])  # 
                    if len(data_list[9]) > 5:
                        # Insert a dash after the 5th character
                        ws.cell(row=start_row + 5, column=3).value = str(data_list[7] + ' ' + data_list[8] + ' ' + data_list[9][:5] + '-' + data_list[9][5:])
                    else:
                        ws.cell(row=start_row + 5, column=3).value = str(data_list[7] + ' ' + data_list[8] + ' ' + data_list[9])  # 
                    ws.cell(row=start_row + 5, column=5).value = data_list[14]  # 
                    
                    ws.cell(row=start_row + 5, column=7).value = f"${self.ADDRESS_CORRECTION_FEE:.2f}"

                    for x in range(2):  # Rows 0 and 1
                        for y in range(7):  # Columns 0 to 6
                            cell = ws.cell(row=start_row + x, column=y + 1)
                            cell.fill = gray_fill
                
                if index == len(data_lists) - 1:

                    upper_border = Border(top=Side(style='thick'))
                    for x in range(7):
                        cell = ws.cell(row=start_row+7, column=x+1)
                        cell.fill = gray_fill
                        cell.border = upper_border
                    

                    totalCell =  ws.cell(row=start_row+7, column=1)
                    totalText = 'Total'
                    totalCell.value = totalText
                    totalCell.font = Font(bold=True)

                    ws.cell(row=start_row+7, column=7).value = f"${total :.2f}"
                    ws.cell(row=4, column=5).value = f"${total :.2f}"
                    ws.cell(row=4, column=5).font = Font(bold=True)

            ws.cell(row=4, column=3).value = project_name

            results_dir = os.path.join(settings.MEDIA_ROOT, 'results')
            os.makedirs(results_dir, exist_ok=True)
            save_path = os.path.join(results_dir, f"{project_name}_Invoice.xlsx")
            wb.save(save_path)
            print(f"Workbook saved for {project_name} at {save_path}")

    def clear_results_directory(self, directory):
        # Check if the directory exists
        if os.path.exists(directory):
            # List all files and directories in the "results" directory
            for filename in os.listdir(directory):
                file_path = os.path.join(directory, filename)
                try:
                    # If it's a file, remove it directly
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)
                    # If it's a directory, remove it and all its contents
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                except Exception as e:
                    print(f'Failed to delete {file_path}. Reason: {e}')

    def delete_non_xlsx_files(self, directory):
        # Loop through all files in the specified directory
        for filename in os.listdir(directory):
            # Construct the full path of the file
            file_path = os.path.join(directory, filename)
            # Check if it's a file (not a directory)
            if os.path.isfile(file_path):
                # If the file does not end with .xlsx, delete it
                if not filename.endswith('.xlsx'):
                    os.remove(file_path)
                    print(f"Deleted: {file_path}")